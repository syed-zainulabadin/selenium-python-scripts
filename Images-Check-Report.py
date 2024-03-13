import time
from selenium import webdriver
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

urls = [
    "https://www.google.com",
    "---"
]

workbook = Workbook()
sheet = workbook.active

browser = webdriver.Firefox()
browser.maximize_window()
time.sleep(2)

row = 1

for url in urls:
    sheet.cell(row=row, column=1, value=url)
    row += 1

    browser.get(url)
    print(url)
    response = requests.get(url)

    scroll_step = 500
    total_steps = 20

    for i in range(total_steps):
        current_scroll_position = i * scroll_step

        browser.execute_script(f"window.scrollTo(0, {current_scroll_position});")
        time.sleep(0.5)

    soup = BeautifulSoup(browser.page_source, "html.parser")

    img_links = soup.find_all("img", src=lambda src: src and src.startswith("https:"))

    for link in img_links:
        img_src = link.get("src")
        # print(img_src)
        sheet.cell(row=row, column=2, value=img_src)
        row += 1
    time.sleep(2)

workbook.save("Images-output.xlsx")

browser.quit()
