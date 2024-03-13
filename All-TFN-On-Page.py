import time
from selenium import webdriver
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

urls = [
    "https://www.google.com",
    "---"
]

workbook = Workbook()
sheet = workbook.active

browser = webdriver.Firefox()
browser.maximize_window()
time.sleep(2)

row = 1

for url in urls:
    sheet.cell(row=row, column=1, value=url)
    row += 1

    browser.get(url)
    print(url)
    response = requests.get(url)

    soup = BeautifulSoup(response.text, "html.parser")

    tel_links = soup.find_all("a", href=lambda href: href and href.startswith("tel:"))

    values = [link["href"][4:] for link in tel_links]

    for value in values:
        sheet.cell(row=row, column=1, value=value)
        row += 1
    time.sleep(2)

workbook.save("all-tfn-output.xlsx")

browser.quit()
